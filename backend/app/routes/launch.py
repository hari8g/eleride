from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
from typing import List, Dict, Any
import json
import subprocess
from pathlib import Path
import pandas as pd
import math
from datetime import datetime, timedelta
from openpyxl import load_workbook


router = APIRouter(prefix="/launch", tags=["launch"])

LAUNCH_XLS = Path("/data/New launch store.xlsx")
STORES_JSON = Path("/artifacts/launch_stores.json")
PLANS_JSON = Path("/artifacts/launch_plans.json")


ALIASES: Dict[str, list[str]] = {
    "store": ["store","store_name","outlet","location_name"],
    "city": ["city","town"],
    "opening_date": ["opening_date","launch_date","go_live","golive_date"],
    "expected_orders_day": ["expected_orders_day","expected_orders/day","expected_orders","orders_day","orders/day","daily_order_target"],
    "peak_hours": ["peak_hours","peak","peak_hrs","peak_window"],
    "address": ["address","store_address","addr"],
    "sla_target_min": ["sla_target_min","sla","sla_target(min)","sla_target"],
    "buffer_riders": ["buffer_riders","buffer","buffer_%","buffer%"],
    "target_orders_per_rider": ["target_orders_per_rider","orders_per_rider","rider_productivity"],
    "avg_km_per_order": ["avg_km_per_order","avg_distance_per_order","distance_per_order_km"],
    "energy_source": ["energy_source","energy","power"],
    "inr_per_order": ["inr_per_order","revenue_per_order","payout_per_order"],
}


def _read_launch_df() -> pd.DataFrame:
    if not LAUNCH_XLS.exists():
        raise HTTPException(status_code=404, detail="new launch store sheet not found")
    # read all sheets and concatenate
    try:
        sheets = pd.read_excel(LAUNCH_XLS, sheet_name=None, engine="openpyxl")
        frames = []
        for name, sdf in (sheets or {}).items():
            if isinstance(sdf, pd.DataFrame) and not sdf.empty:
                frames.append(sdf)
        if not frames:
            raise ValueError("empty_sheets")
        df = pd.concat(frames, ignore_index=True)
    except Exception:
        # Fallback: manual parse via openpyxl to avoid ambiguous pandas typing
        wb = load_workbook(LAUNCH_XLS, read_only=True, data_only=True)
        rows_acc: list[dict[str, Any]] = []
        for ws in wb.worksheets:
            headers: list[str] = []
            for ridx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if not headers:
                    # find first non-empty row as headers
                    if row and any(c is not None and str(c).strip() != "" for c in row):
                        headers = [str(c).strip() if c is not None else "" for c in row]
                    continue
                if not row or all(c is None or str(c).strip()=="" for c in row):
                    continue
                rec = {}
                for i, h in enumerate(headers):
                    if not h:
                        continue
                    rec[h] = row[i] if i < len(row) else None
                rows_acc.append(rec)
        if not rows_acc:
            raise HTTPException(status_code=400, detail="launch sheet has no data (fallback)")
        df = pd.DataFrame(rows_acc)
    # normalize columns and apply aliases
    cols_norm = [str(c).strip().lower().replace(" ", "_") for c in df.columns]
    df.columns = cols_norm
    # drop duplicate columns (keep first) to avoid DataFrame returns on df["col"]
    try:
        df = df.loc[:, ~pd.Index(df.columns).duplicated(keep='first')]
    except Exception:
        pass
    # build a mapping from alias -> canonical
    colmap: Dict[str, str] = {}
    for canon, als in ALIASES.items():
        for a in als:
            if a in df.columns:
                colmap[a] = canon
    # heuristic mapping for unknown headers by substring
    for c in list(df.columns):
        if c in colmap:
            continue
        cn = None
        if "store" in c:
            # Avoid mapping Store Type to store id
            if "type" in c:
                cn = None
            elif any(k in c for k in ["name","outlet","location"]):
                cn = "store"
        elif "city" in c:
            cn = "city"
        elif ("open" in c and "date" in c) or ("launch" in c and "date" in c) or ("go_live" in c):
            cn = "opening_date"
        elif ("order" in c and ("day" in c or "/day" in c)) or ("daily" in c and "order" in c):
            cn = "expected_orders_day"
        elif "sla" in c:
            cn = "sla_target_min"
        elif "buffer" in c:
            cn = "buffer_riders"
        elif ("rider" in c and ("target" in c or "per" in c)):
            cn = "target_orders_per_rider"
        elif ("km" in c and "order" in c) or ("distance" in c and "order" in c):
            cn = "avg_km_per_order"
        elif "energy" in c:
            cn = "energy_source"
        elif ("inr" in c and "order" in c) or ("payout" in c and "order" in c) or ("revenue" in c and "order" in c):
            cn = "inr_per_order"
        elif "peak" in c:
            cn = "peak_hours"
        elif "address" in c or "addr" in c:
            cn = "address"
        if cn:
            colmap[c] = cn
    # rename known aliases to canonical
    df = df.rename(columns=colmap)
    # drop duplicates after renaming as well
    try:
        df = df.loc[:, ~pd.Index(df.columns).duplicated(keep='first')]
    except Exception:
        pass
    for c in ["expected_orders_day","sla_target_min","buffer_riders","target_orders_per_rider","daily_order_target"]:
        if c in df.columns:
            try:
                series = df[c]
                if isinstance(series, pd.DataFrame):
                    series = series.iloc[:,0]
                df[c] = pd.to_numeric(series, errors="coerce")
            except Exception:
                pass
    # if expected_orders_day missing, derive from daily_order_target
    if "expected_orders_day" not in df.columns and "daily_order_target" in df.columns:
        df["expected_orders_day"] = df["daily_order_target"]
    # coerce dates
    if "opening_date" in df.columns:
        series = df["opening_date"]
        if isinstance(series, pd.DataFrame):
            series = series.iloc[:,0]
        df["opening_date"] = pd.to_datetime(series, errors="coerce")
    # drop rows fully empty
    df = df.dropna(how='all')
    # ensure we have a 'store' column; synthesize if missing
    if 'store' not in df.columns:
        # try create from first non-empty among likely columns per row
        likely_cols = [c for c in df.columns if any(k in c for k in ["store","outlet","location","name"])]
        if likely_cols:
            df['store'] = df[likely_cols].astype(str).apply(
                lambda s: next((v for v in s.values if str(v).strip().lower() not in ('', 'nan', 'none')), None),
                axis=1
            )
    # trim and drop rows without store (fallback to any store-like col per row)
    if 'store' in df.columns:
        s = df['store']
        if isinstance(s, pd.DataFrame):
            s = s.iloc[:,0]
        df['store'] = s.astype(str).apply(lambda x: x.strip()).replace({'': None, 'nan': None, 'None': None})
    else:
        df['store'] = None
    store_like = [c for c in df.columns if any(k in c for k in ["store","store_name","outlet","location","name"]) and c != 'store']
    def _row_store(r):
        if r.get('store'):
            return r.get('store')
        for c in store_like:
            v = r.get(c)
            if v is None:
                continue
            vs = str(v).strip()
            if vs.lower() not in ('', 'nan', 'none'):
                return vs
        return None
    df['store'] = df.apply(_row_store, axis=1)
    df = df[df['store'].notna()]

    # if sheet is slabbed by store (multiple rows per store), aggregate to one row per store
    if 'expected_orders_day' not in df.columns:
        # map from any alias again
        for cand in [
            'daily_order_target','orders_day','orders/day','expected_orders','expected_orders/day'
        ]:
            if cand in df.columns:
                df['expected_orders_day'] = pd.to_numeric(df[cand], errors='coerce')
                break

    group_fields = {
        'city': 'first',
        'opening_date': 'first',
        'expected_orders_day': 'max',
        'sla_target_min': 'first',
        'address': 'first',
        'peak_hours': 'first',
        'buffer_riders': 'first',
        'target_orders_per_rider': 'first',
        'avg_km_per_order': 'first',
        'energy_source': 'first',
        'inr_per_order': 'first',
    }
    agg_cols = {k:v for k,v in group_fields.items() if k in df.columns}
    if agg_cols:
        df = df.groupby('store', as_index=False).agg(agg_cols)

    return df.reset_index(drop=True)


class LaunchStore(BaseModel):
    store: str
    city: str | None = None
    opening_date: str | None = None
    readiness_score: float
    risk: str | None = None


@router.get("/stores", response_model=List[LaunchStore])
def list_launch_stores(debug: bool = False) -> List[LaunchStore]:
    # Prefer preprocessed artifacts if available
    if STORES_JSON.exists():
        try:
            data = json.loads(STORES_JSON.read_text())
            return [LaunchStore(**{
                "store": d.get("store"),
                "city": d.get("city"),
                "opening_date": d.get("opening_date"),
                "readiness_score": float(d.get("readiness_score") or 0),
                "risk": d.get("risk"),
            }) for d in data]
        except Exception:
            pass
    try:
        df = _read_launch_df()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"launch_read_error: {e}")
    if debug:
        # return minimal debug info in exception to avoid changing response model
        cols = list(df.columns)
        sample = df.head(3).to_dict(orient='records')
        raise HTTPException(status_code=200, detail={"columns": cols, "sample": sample})
    # group by store to consolidate multiple slab rows
    out: List[LaunchStore] = []
    if 'store' not in df.columns:
        return out
    g = df.groupby('store', dropna=True)
    for store, grp in g:
        if not store:
            continue
        city = None
        if 'city' in grp.columns:
            city = next((str(x).strip() for x in grp['city'].tolist() if pd.notna(x) and str(x).strip()), None)
        open_date = None
        if 'opening_date' in grp.columns:
            try:
                od = next((x for x in grp['opening_date'].tolist() if pd.notna(x)), None)
                open_date = od.strftime('%Y-%m-%d') if od is not None else None
            except Exception:
                open_date = None
        fields = ["expected_orders_day","opening_date","sla_target_min","address","peak_hours"]
        have = 0
        for c in fields:
            if c in grp.columns and grp[c].notna().any():
                have += 1
        score = (have/len(fields))*60.0
        if 'expected_orders_day' in grp.columns and grp['expected_orders_day'].notna().any():
            score += 25.0
        if 'energy_source' in grp.columns and grp['energy_source'].astype(str).apply(lambda x: x.strip()).replace({'': None, 'nan': None, 'None': None}).notna().any():
            score += 15.0
        score = max(0.0, min(100.0, score))
        risk = None
        if score < 50:
            risk = 'incomplete data'
        elif score < 70:
            risk = 'needs staffing/energy check'
        out.append(LaunchStore(store=str(store), city=city, opening_date=open_date, readiness_score=round(score,1), risk=risk))
    return out


@router.get("/{store}/plan")
def launch_plan(store: str) -> Dict[str, Any]:
    # Prefer preprocessed plan artifact
    if PLANS_JSON.exists():
        try:
            plans = json.loads(PLANS_JSON.read_text())
            if store in plans:
                return plans[store]
        except Exception:
            pass
    df = _read_launch_df()
    row = None
    for _, r in df.iterrows():
        if str(r.get("store") or r.get("store_name") or "").strip() == store:
            row = r
            break
    if row is None:
        raise HTTPException(status_code=404, detail="store not found in launch sheet")

    expected_orders_day = float(row.get("expected_orders_day") or 0)
    target_orders_per_rider = float(row.get("target_orders_per_rider") or 22)
    buffer_pct = float(row.get("buffer_riders") or 15) / 100.0
    riders_day = expected_orders_day / max(1.0, target_orders_per_rider)
    riders_day_with_buffer = riders_day * (1.0 + buffer_pct)

    # shift breakdown heuristics
    peak = (str(row.get("peak_hours") or "").lower())
    if any(k in peak for k in ["morning","am","11-14","10-14"]):
        morning_ratio = 0.6
    else:
        morning_ratio = 0.5
    morning = math.ceil(riders_day_with_buffer * morning_ratio)
    evening = math.ceil(riders_day_with_buffer - morning)

    # energy estimation
    avg_km_per_order = float(row.get("avg_km_per_order") or 2.5)
    kwh_per_km = 0.03
    energy_kwh_day = expected_orders_day * avg_km_per_order * kwh_per_km
    battery_kwh = 2.0
    swaps_day = energy_kwh_day / battery_kwh if battery_kwh > 0 else 0.0

    # SLA feasibility (very rough):
    sla_target = float(row.get("sla_target_min") or 30)
    headroom = max(0.0, (riders_day_with_buffer - riders_day) / max(1.0, riders_day))
    predicted_sla = max(15.0, sla_target - headroom*10.0)

    # 4-week ROI curve using simple ramp
    inr_per_order = float(row.get("inr_per_order") or 200)
    week0 = expected_orders_day * inr_per_order * 6.5
    w1 = week0 * 0.9
    w2 = week0 * 1.0
    w3 = week0 * 1.1
    w4 = week0 * 1.15

    plan = {
        "store": store,
        "city": str(row.get("city") or "").strip() or None,
        "opening_date": (row.get("opening_date").strftime("%Y-%m-%d") if pd.notna(row.get("opening_date")) else None),
        "staffing": {
            "riders_per_day": math.ceil(riders_day_with_buffer),
            "target_orders_per_rider": target_orders_per_rider,
            "buffer_pct": round(buffer_pct*100,1),
            "shifts": [
                {"name":"Morning","riders": morning},
                {"name":"Evening","riders": evening},
            ],
        },
        "energy": {
            "avg_km_per_order": avg_km_per_order,
            "kwh_per_km": kwh_per_km,
            "energy_kwh_day": round(energy_kwh_day,2),
            "swaps_day": round(swaps_day,1),
        },
        "sla": {
            "target_min": sla_target,
            "predicted_min": round(predicted_sla,1),
        },
        "roi": {
            "weekly_inr": [round(v,2) for v in [w1,w2,w3,w4]],
            "assumptions": {"inr_per_order": inr_per_order}
        },
    }
    return plan


class LaunchTask(BaseModel):
    task: str
    owner: str
    due: str
    status: str


@router.get("/{store}/tasks", response_model=List[LaunchTask])
def launch_tasks(store: str) -> List[LaunchTask]:
    # If plan exists in artifacts, keep tasks as before (derived from opening_date)
    df = _read_launch_df()
    row = None
    for _, r in df.iterrows():
        if str(r.get("store") or r.get("store_name") or "").strip() == store:
            row = r
            break
    if row is None:
        raise HTTPException(status_code=404, detail="store not found in launch sheet")
    open_date = row.get("opening_date")
    if pd.isna(open_date):
        base = datetime.utcnow().date()
    else:
        base = open_date.date()
    tasks = [
        ("Permits & compliance", "Ops", base - timedelta(days=14), "pending"),
        ("Station/energy setup", "Ops", base - timedelta(days=10), "pending"),
        ("3PL partner onboarding", "BizDev", base - timedelta(days=7), "pending"),
        ("Rider hiring/training", "Ops", base - timedelta(days=5), "pending"),
        ("Soft launch", "Ops", base - timedelta(days=2), "pending"),
        ("Go-live", "Ops", base, "pending"),
    ]
    return [LaunchTask(task=t, owner=o, due=d.strftime("%Y-%m-%d"), status=s) for (t,o,d,s) in tasks]


@router.get("/debug")
def launch_debug() -> Dict[str, Any]:
    if not LAUNCH_XLS.exists():
        raise HTTPException(status_code=404, detail="new launch store sheet not found")
    info: Dict[str, Any] = {"file": str(LAUNCH_XLS), "sheets": []}
    try:
        wb = load_workbook(LAUNCH_XLS, read_only=True, data_only=True)
        for ws in wb.worksheets:
            sheet: Dict[str, Any] = {"name": ws.title}
            headers = []
            rows = []
            for ridx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                vals = [(str(c).strip() if c is not None else None) for c in row]
                if not headers and any(v for v in vals):
                    headers = vals
                    continue
                if headers and any(v for v in vals):
                    rows.append(vals)
                if len(rows) >= 5:
                    break
            sheet["headers"] = headers
            sheet["sample_rows"] = rows
            info["sheets"].append(sheet)
    except Exception as e:
        info["error"] = f"openpyxl_error: {e}"
    # also try our DataFrame view
    try:
        df = _read_launch_df()
        info["normalized_columns"] = list(df.columns)
        info["normalized_sample"] = df.head(5).to_dict(orient='records')
    except Exception as e:
        info["normalized_error"] = str(e)
    return info


@router.post("/reprocess")
def launch_reprocess() -> Dict[str, Any]:
    """Run the launch preprocessing script to rebuild artifacts."""
    try:
        # Prefer calling the script directly so it can evolve independently
        res = subprocess.run(["python", "/analytics/preprocess_launch.py"], capture_output=True, text=True, check=True)
        return {"status": "ok", "stdout": res.stdout[-4000:], "stderr": res.stderr[-4000:]}
    except subprocess.CalledProcessError as e:
        raise HTTPException(status_code=500, detail={"error": "preprocess_failed", "stdout": e.stdout[-4000:] if e.stdout else "", "stderr": e.stderr[-4000:] if e.stderr else ""})

